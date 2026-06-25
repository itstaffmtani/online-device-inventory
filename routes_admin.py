# routes_admin.py — blueprint admin (dashboard 3 tab, dua detail, export PDF+XLSX).
#
#   GET  /admin                       -> dashboard: 3 tab (Laptop/Karyawan/Pengadaan)
#   GET  /admin/laptop/<device_id>    -> detail + riwayat 1 laptop
#   GET  /admin/device/<id>           -> redirect 302 ke /admin/laptop/<id> (back-compat)
#   GET  /admin/karyawan/<employee_id>-> detail + riwayat 1 karyawan
#   GET  /admin/laptop/<id>/export.pdf    -> PDF laporan 1 laptop
#   GET  /admin/karyawan/<id>/export.pdf  -> PDF laporan 1 karyawan
#   GET  /admin/export.xlsx           -> export terbaru (sheet Laptop + Per Karyawan)
#   GET/POST /admin/login             -> login 1 password bersama (ADMIN_PASSWORD)
#   GET  /admin/logout                -> keluar
#
# Auth: 1 password bersama (env ADMIN_PASSWORD) + Flask session. Tanpa tabel user.

import io
import json
from datetime import datetime
from functools import wraps

from flask import (Blueprint, Response, current_app, flash, redirect,
                   render_template, request, session, url_for)

import db
import scoring_config
from scoring import (build_insights, cpu_passmark, get_profiles,
                     score_submission, suggest_placement, PASSMARK_PER_THREAD)

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def wg_labels():
    """Peta {key: label} kelompok kerja (data-driven dari tabel work_groups)."""
    return scoring_config.get_labels()


def group_label(row):
    """Label kelompok kerja untuk tampilan; pakai teks bebas bila 'Lainnya'."""
    wg = row.get("work_group")
    if wg == "other" and (row.get("work_group_other") or "").strip():
        return row["work_group_other"].strip()
    return wg_labels().get(wg, wg or "-")
STATUS_LABEL = {"eligible": "Layak", "upgrade": "Upgrade", "replace": "Ganti"}
LAPTOP_STATUS_LABEL = {"office_inventory": "Inventaris Kantor", "personal": "Milik Pribadi"}
PHYSICAL_CONDITION_LABEL = {"good": "Baik", "fair": "Cukup", "poor": "Kurang"}
SOURCE_LABEL = {
    "windows_script": "Script Windows", "mac_script": "Script Mac",
    "linux_script": "Script Linux", "manual": "Manual",
}
_STATUS_RANK = {"replace": 0, "upgrade": 1, "eligible": 2}


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin"):
            # Simpan path LENGKAP (termasuk prefix sub-path) agar redirect balik benar.
            return redirect(url_for("admin.login", next=request.script_root + request.path))
        return view(*args, **kwargs)
    return wrapped


def _parse_reasons(raw):
    if not raw:
        return []
    try:
        val = json.loads(raw)
        return val if isinstance(val, list) else [str(val)]
    except (ValueError, TypeError):
        return [raw]


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == current_app.config["ADMIN_PASSWORD"]:
            session["admin"] = True
            nxt = request.args.get("next") or ""
            # Hanya izinkan path lokal (cegah open-redirect ke domain luar).
            if not nxt.startswith("/") or nxt.startswith("//"):
                nxt = url_for("admin.dashboard")
            return redirect(nxt)
        error = "Password salah."
    return render_template("admin/login.html", error=error)


@admin_bp.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect(url_for("admin.login"))


# ---------------------------------------------------------------------------
# Dashboard + list
# ---------------------------------------------------------------------------
@admin_bp.route("/")
@login_required
def dashboard():
    rows = db.latest_per_device()
    current_year = datetime.now().year
    for r in rows:
        r["_reasons"] = _parse_reasons(r.get("status_reasons"))
        r["_insight"] = build_insights(r, current_year=current_year)
        r["_suggest"] = suggest_placement(r, current_year=current_year)

    # Ringkasan.
    summary = {
        "total": len(rows),
        "per_status": {"eligible": 0, "upgrade": 0, "replace": 0},
        "per_group": {},
        "per_company": {},
    }
    # Agregat insight: masalah paling umum di seluruh armada laptop.
    attention = {"need_action": 0, "ram_kurang": 0, "belum_ssd": 0, "baterai_lemah": 0}
    for r in rows:
        st = r.get("status")
        if st in summary["per_status"]:
            summary["per_status"][st] += 1
        grp = r.get("work_group") or "-"
        summary["per_group"][grp] = summary["per_group"].get(grp, 0) + 1
        comp = r.get("holder_company") or "-"
        summary["per_company"][comp] = summary["per_company"].get(comp, 0) + 1
        # Hitung masalah umum dari komponen insight.
        if st in ("upgrade", "replace"):
            attention["need_action"] += 1
        for c in r["_insight"]["components"]:
            if c["tone"] != "bad":
                continue
            if c["label"] == "RAM":
                attention["ram_kurang"] += 1
            elif c["label"] == "Penyimpanan":
                attention["belum_ssd"] += 1
            elif c["label"] == "Baterai":
                attention["baterai_lemah"] += 1

    # — TAB Pengadaan: agregat keputusan (dihitung dari semua baris armada) —
    pengadaan = _build_pengadaan(rows, current_year)

    # Cari (hanya memengaruhi tabel Laptop).
    q = (request.args.get("q") or "").strip().lower()
    if q:
        def match(r):
            hay = " ".join(str(r.get(k) or "") for k in (
                "holder_name", "serial_number", "device_serial", "device_brand",
                "device_model", "work_group", "holder_company", "asset_no"))
            return q in hay.lower()
        rows = [r for r in rows if match(r)]

    # Sortir.
    sort = request.args.get("sort", "last_seen")
    direction = request.args.get("dir", "desc")
    keyfuncs = {
        "name": lambda r: (r.get("holder_name") or "").lower(),
        "serial": lambda r: (r.get("serial_number") or r.get("device_serial") or "").lower(),
        "brand": lambda r: (r.get("device_brand") or "").lower(),
        "status": lambda r: _STATUS_RANK.get(r.get("status"), -1),
        "score": lambda r: r.get("score_total") if r.get("score_total") is not None else -1,
        "group": lambda r: r.get("work_group") or "",
        "last_seen": lambda r: r.get("submitted_at") or "",
    }
    keyf = keyfuncs.get(sort, keyfuncs["last_seen"])
    rows.sort(key=keyf, reverse=(direction == "desc"))

    # — TAB Karyawan: satu baris per karyawan = submission terbaru miliknya —
    employees = []
    try:
        employees = db.latest_per_employee()
    except AttributeError:
        # db.latest_per_employee belum tersedia (migrasi belum jalan) — tab kosong.
        employees = []

    # Tab aktif (?view=laptop|karyawan|pengadaan).
    view = request.args.get("view", "laptop")
    if view not in ("laptop", "karyawan", "pengadaan"):
        view = "laptop"

    return render_template("admin/dashboard.html", rows=rows, summary=summary,
                           attention=attention, pengadaan=pengadaan,
                           employees=employees, view=view, current_year=current_year,
                           q=request.args.get("q", ""), sort=sort, dir=direction,
                           wg_label=wg_labels(), group_label=group_label,
                           status_label=STATUS_LABEL)


def _months_since(ts):
    """Berapa bulan (perkiraan) sejak timestamp ISO; None bila tak terbaca."""
    if not ts:
        return None
    try:
        dt = datetime.fromisoformat(str(ts).replace("Z", "").split(".")[0])
    except (ValueError, TypeError):
        return None
    now = datetime.now()
    return (now.year - dt.year) * 12 + (now.month - dt.month)


def _build_pengadaan(rows, current_year):
    """Agregat keputusan pengadaan (kartu actionable) dari latest_per_device().

    Tiap kategori menyimpan daftar singkat unit (nama + laptop) agar bisa
    ditindaklanjuti, bukan sekadar angka.
    """
    def unit(r):
        return {
            "device_id": r.get("device_id"),
            "holder_name": r.get("holder_name") or "-",
            "laptop": ((r.get("device_brand") or "-") + " "
                       + (r.get("device_model") or "")).strip(),
            "serial": r.get("serial_number") or r.get("device_serial") or "-",
            "group_label": group_label(r),
            "detail": "",
        }

    perlu_ganti, mendekati_pensiun, belum_win11, disk_lemah, kedaluwarsa = [], [], [], [], []
    per_group = {}  # kelompok -> jumlah perlu tindakan (upgrade+replace)

    for r in rows:
        st = r.get("status")
        if st in ("upgrade", "replace"):
            grp = group_label(r)
            per_group[grp] = per_group.get(grp, 0) + 1
        if st == "replace":
            perlu_ganti.append(unit(r))
        eol = r.get("eol_year")
        if eol is not None and eol <= current_year + 1:
            u = unit(r); u["detail"] = f"Pensiun {eol}"; mendekati_pensiun.append(u)
        # Lewati bila OS sudah Windows 11 (flag tak relevan; cegah false negative
        # collector tanpa admin yang gagal baca TPM/Secure Boot).
        os_name = (r.get("os_name") or "").lower()
        already_win11 = "windows" in os_name and "11" in os_name
        if r.get("win11_ready") in (0, "0") and not already_win11:
            u = unit(r)
            u["detail"] = (r.get("win11_blockers") or "Belum memenuhi syarat").strip()
            belum_win11.append(u)
        dh = r.get("disk_health_pct")
        if dh is not None and dh < 50:
            u = unit(r); u["detail"] = f"Disk {round(dh)}%"; disk_lemah.append(u)
        months = _months_since(r.get("submitted_at"))
        if months is not None and months > 6:
            u = unit(r); u["detail"] = f"{months} bln lalu"; kedaluwarsa.append(u)

    return {
        "perlu_ganti": perlu_ganti,
        "mendekati_pensiun": mendekati_pensiun,
        "belum_win11": belum_win11,
        "disk_lemah": disk_lemah,
        "kedaluwarsa": kedaluwarsa,
        "per_group": per_group,
    }


@admin_bp.route("/laptop/<int:device_id>")
@login_required
def laptop_detail(device_id):
    data = db.device_with_history(device_id)
    if not data.get("device"):
        return Response("Laptop tidak ditemukan.", status=404)
    for s in data["submissions"]:
        s["_reasons"] = _parse_reasons(s.get("status_reasons"))
    latest = data["submissions"][0] if data["submissions"] else None
    current_year = datetime.now().year
    insights = build_insights(latest, current_year=current_year) if latest else None
    suggest = suggest_placement(latest, current_year=current_year) if latest else None
    return render_template("admin/detail.html", device=data["device"],
                           submissions=data["submissions"], latest=latest,
                           wg_label=wg_labels(), group_label=group_label,
                           status_label=STATUS_LABEL, profiles=get_profiles(),
                           insights=insights, suggest=suggest)


@admin_bp.route("/device/<int:device_id>")
@login_required
def device_detail(device_id):
    # Back-compat: slug lama -> redirect 302 ke /admin/laptop/<id>.
    return redirect(url_for("admin.laptop_detail", device_id=device_id), code=302)


@admin_bp.route("/karyawan/<int:employee_id>")
@login_required
def employee_detail(employee_id):
    data = db.employee_with_history(employee_id)
    if not data.get("employee"):
        return Response("Karyawan tidak ditemukan.", status=404)
    for s in data["submissions"]:
        s["_reasons"] = _parse_reasons(s.get("status_reasons"))
    latest = data["submissions"][0] if data["submissions"] else None
    current_year = datetime.now().year
    insights = build_insights(latest, current_year=current_year) if latest else None
    suggest = suggest_placement(latest, current_year=current_year) if latest else None
    return render_template("admin/employee_detail.html", employee=data["employee"],
                           submissions=data["submissions"], latest=latest,
                           wg_label=wg_labels(), group_label=group_label,
                           status_label=STATUS_LABEL, profiles=get_profiles(),
                           insights=insights, suggest=suggest)


# ---------------------------------------------------------------------------
# EDIT submission (admin-only) + Hitung ulang skor (per laptop & massal)
# ---------------------------------------------------------------------------
# Kolom yang boleh diedit admin, dikelompokkan untuk render form (label, kind).
# kind: "text" | "number" | "textarea" | "group" | list[(value,label)] (select)
_LAPTOP_STATUS_OPTS = [("", "—"), ("office_inventory", "Inventaris Kantor"),
                       ("personal", "Milik Pribadi")]
_PHYS_OPTS = [("", "—"), ("good", "Baik"), ("fair", "Cukup"), ("poor", "Kurang")]
_YESNO_OPTS = [("", "Tak diketahui"), ("1", "Ya"), ("0", "Tidak")]

EDIT_SECTIONS = [
    ("Pemegang & Penempatan", [
        ("holder_name", "Nama", "text"),
        ("holder_position", "Jabatan", "text"),
        ("holder_company", "Perusahaan", "text"),
        ("holder_location", "Lokasi/Cabang", "text"),
        ("work_group", "Kelompok Kerja", "group"),
        ("work_group_other", "Kelompok (bila Lainnya)", "text"),
        ("laptop_status", "Status Kepemilikan", _LAPTOP_STATUS_OPTS),
    ]),
    ("Identitas Aset", [
        ("serial_number", "Nomor Seri", "text"),
        ("asset_no", "No. Aset", "text"),
        ("hostname", "Hostname", "text"),
        ("mac_address", "MAC Address", "text"),
    ]),
    ("CPU & GPU", [
        ("cpu_model", "CPU / Prosesor", "text"),
        ("cpu_cores", "Core CPU", "number"),
        ("cpu_threads", "Thread CPU", "number"),
        ("cpu_speed_mhz", "Kecepatan CPU (MHz)", "number"),
        ("gpu", "GPU / Kartu Grafis", "text"),
        ("motherboard", "Motherboard", "text"),
    ]),
    ("RAM", [
        ("ram_gb", "RAM (GB)", "number"),
        ("ram_type", "Tipe RAM", "text"),
        ("ram_speed_mhz", "Kecepatan RAM (MHz)", "number"),
        ("ram_usage_pct", "Pemakaian RAM saat cek (%)", "number"),
        ("ram_slots_total", "Slot RAM Total", "number"),
        ("ram_slots_used", "Slot RAM Terisi", "number"),
        ("ram_max_gb", "RAM Maksimum (GB)", "number"),
    ]),
    ("Penyimpanan & OS", [
        ("ssd_gb", "SSD (GB)", "number"),
        ("ssd_type", "Tipe SSD (NVMe/SATA)", "text"),
        ("hdd_gb", "HDD (GB)", "number"),
        ("os_total_gb", "Total Disk OS (GB)", "number"),
        ("os_free_gb", "Sisa Disk OS (GB)", "number"),
        ("disk_health_pct", "Kesehatan Disk (%)", "number"),
        ("os_name", "Sistem Operasi", "text"),
    ]),
    ("Windows 11 & Keamanan", [
        ("tpm_version", "Versi TPM", "text"),
        ("secure_boot", "Secure Boot", _YESNO_OPTS),
        ("win11_ready", "Siap Windows 11 (indikasi)", _YESNO_OPTS),
        ("win11_blockers", "Kendala Windows 11", "text"),
    ]),
    ("Baterai", [
        ("battery_pct", "Daya saat cek (%)", "number"),
        ("battery_health_pct", "Kesehatan Baterai (%)", "number"),
        ("battery_wh_full", "Kapasitas Penuh (Wh)", "number"),
        ("battery_wh_design", "Kapasitas Desain (Wh)", "number"),
    ]),
    ("Kondisi & Kelengkapan", [
        ("physical_condition", "Kondisi Fisik", _PHYS_OPTS),
        ("accessories", "Kelengkapan", "text"),
        ("purchase_year", "Tahun Pembelian", "number"),
        ("issues", "Keluhan / Kerusakan", "textarea"),
    ]),
]

_EDIT_INT_COLS = {"cpu_cores", "cpu_threads", "cpu_speed_mhz", "ram_speed_mhz",
                  "ram_slots_total", "ram_slots_used", "secure_boot", "win11_ready",
                  "purchase_year"}
_EDIT_FLOAT_COLS = {"ram_gb", "ram_usage_pct", "ram_max_gb", "ssd_gb", "hdd_gb",
                    "os_total_gb", "os_free_gb", "disk_health_pct", "battery_pct",
                    "battery_health_pct", "battery_wh_full", "battery_wh_design"}
# Semua kolom yang boleh diset dari form edit.
_EDIT_COLS = [c for _sec, fields in EDIT_SECTIONS for (c, _l, _k) in fields]


def _to_int(v):
    try:
        return int(float(str(v).replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _to_float(v):
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _rescore_fields(sub):
    """Hitung ulang skor untuk dict `sub` (kolom DB) -> dict kolom skor utk update.

    Re-resolve cpu_passmark dari cpu_model agar perubahan CPU ikut terhitung;
    namun hormati nilai cpu_passmark tersimpan bila model tak dikenali.
    """
    s = dict(sub)
    pm, est = cpu_passmark(s.get("cpu_model"))
    if est:
        existing = _to_float(s.get("cpu_passmark"))
        if existing:  # hormati nilai manual/tersimpan yang sudah ada
            pm = int(existing)
            est = False
        else:
            threads = _to_int(s.get("cpu_threads")) or _to_int(s.get("cpu_cores"))
            pm = int(round(threads * PASSMARK_PER_THREAD)) if threads else 0
    s["cpu_passmark"] = pm
    s["cpu_estimated"] = est
    result = score_submission(s, current_year=datetime.now().year)
    out = dict(result)
    out["status_reasons"] = json.dumps(result["status_reasons"], ensure_ascii=False)
    out["cpu_passmark"] = pm
    return out


@admin_bp.route("/submission/<int:submission_id>/edit", methods=["GET", "POST"])
@login_required
def edit_submission(submission_id):
    sub = db.get_submission(submission_id)
    if not sub:
        return Response("Submission tidak ditemukan.", status=404)

    if request.method == "POST":
        valid_keys = {g["key"] for g in scoring_config.all_groups()}
        valid_keys.add("other")
        data = {}
        for col in _EDIT_COLS:
            if col not in request.form:
                continue
            raw = (request.form.get(col) or "").strip()
            if col in _EDIT_INT_COLS:
                data[col] = _to_int(raw) if raw != "" else None
            elif col in _EDIT_FLOAT_COLS:
                data[col] = _to_float(raw) if raw != "" else None
            else:
                data[col] = raw or None

        # Validasi & normalisasi kelompok kerja.
        wg = data.get("work_group")
        if wg and wg not in valid_keys:
            flash("Kelompok kerja tidak valid.", "error")
            return redirect(url_for("admin.edit_submission", submission_id=submission_id))
        if wg != "other":
            data["work_group_other"] = None
        if not (data.get("holder_name") or "").strip():
            flash("Nama pemegang wajib diisi.", "error")
            return redirect(url_for("admin.edit_submission", submission_id=submission_id))

        # Gabung data lama + perubahan, lalu hitung ulang skor.
        merged = {**sub, **data}
        data.update(_rescore_fields(merged))
        db.update_submission(submission_id, data)
        flash("Perubahan tersimpan & skor dihitung ulang.", "ok")
        return redirect(url_for("admin.laptop_detail", device_id=sub["device_id"]))

    return render_template(
        "admin/edit_submission.html", sub=sub, sections=EDIT_SECTIONS,
        groups=scoring_config.all_groups(active_only=False),
        status_label=STATUS_LABEL,
    )


@admin_bp.route("/laptop/<int:device_id>/recalc", methods=["POST"])
@login_required
def recalc_laptop(device_id):
    data = db.device_with_history(device_id)
    if not data.get("device"):
        return Response("Laptop tidak ditemukan.", status=404)
    n = 0
    for s in data["submissions"]:
        db.update_submission(s["id"], _rescore_fields(s))
        n += 1
    flash(f"Skor {n} riwayat laptop ini diperbarui.", "ok")
    return redirect(url_for("admin.laptop_detail", device_id=device_id))


@admin_bp.route("/recalc-all", methods=["POST"])
@login_required
def recalc_all():
    n = 0
    for s in db.all_submissions():
        db.update_submission(s["id"], _rescore_fields(s))
        n += 1
    flash(f"Skor {n} submission diperbarui sesuai parameter terbaru.", "ok")
    return redirect(url_for("admin.dashboard"))


# ---------------------------------------------------------------------------
# Parameter skoring & kelola kelompok kerja (/admin/skoring)
# ---------------------------------------------------------------------------
@admin_bp.route("/skoring")
@login_required
def scoring_page():
    return render_template(
        "admin/scoring.html",
        groups=scoring_config.all_groups(active_only=False),
        settings=scoring_config.get_settings(),
    )


@admin_bp.route("/skoring/grup/<key>", methods=["POST"])
@login_required
def scoring_update_group(key):
    fields = {"label": request.form.get("label")}
    for col in ("cpu_floor", "cpu_ideal", "ram_min", "ram_ideal", "sort_order"):
        fields[col] = _to_int(request.form.get(col))
    for col in ("w_cpu", "w_ram", "w_storage", "w_battery"):
        fields[col] = _to_float(request.form.get(col))
    fields["is_active"] = request.form.get("is_active") == "on"
    scoring_config.update_group(key, fields)
    flash(f"Kelompok '{key}' diperbarui. Klik 'Hitung ulang semua' agar skor ikut berubah.", "ok")
    return redirect(url_for("admin.scoring_page"))


@admin_bp.route("/skoring/grup", methods=["POST"])
@login_required
def scoring_create_group():
    profile = {}
    for col in ("cpu_floor", "cpu_ideal", "ram_min", "ram_ideal", "sort_order"):
        v = _to_int(request.form.get(col))
        if v is not None:
            profile[col] = v
    for col in ("w_cpu", "w_ram", "w_storage", "w_battery"):
        v = _to_float(request.form.get(col))
        if v is not None:
            profile[col] = v
    ok, msg = scoring_config.create_group(
        request.form.get("key"), request.form.get("label"), profile)
    flash(("Kelompok baru ditambahkan." if ok else msg), "ok" if ok else "error")
    return redirect(url_for("admin.scoring_page"))


@admin_bp.route("/skoring/ambang", methods=["POST"])
@login_required
def scoring_update_settings():
    fields = {}
    for k in ("status_eligible_min", "status_upgrade_min", "base_lifespan_years",
              "blend_spec", "blend_load"):
        v = _to_float(request.form.get(k))
        if v is not None:
            fields[k] = v
    scoring_config.update_settings(fields)
    flash("Ambang skoring diperbarui. Klik 'Hitung ulang semua' agar skor ikut berubah.", "ok")
    return redirect(url_for("admin.scoring_page"))


# ---------------------------------------------------------------------------
# Export PDF (fpdf2) — laporan per laptop & per karyawan.
# ---------------------------------------------------------------------------
def _ascii_filename(text, fallback):
    """Slug nama file ASCII aman untuk header Content-Disposition."""
    import re
    s = (text or "").strip()
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9._-]+", "-", s).strip("-")
    return s or fallback


@admin_bp.route("/laptop/<int:device_id>/export.pdf")
@login_required
def laptop_export_pdf(device_id):
    import pdf_export
    data = db.device_with_history(device_id)
    if not data.get("device"):
        return Response("Laptop tidak ditemukan.", status=404)
    latest = data["submissions"][0] if data["submissions"] else None
    insights = build_insights(latest, current_year=datetime.now().year) if latest else None
    pdf_bytes = pdf_export.laptop_pdf(data["device"], data["submissions"], latest, insights)
    serial = data["device"].get("serial_number") or f"id{device_id}"
    fname = f"laptop-{_ascii_filename(serial, 'laptop')}.pdf"
    return Response(pdf_bytes, mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@admin_bp.route("/karyawan/<int:employee_id>/export.pdf")
@login_required
def employee_export_pdf(employee_id):
    import pdf_export
    data = db.employee_with_history(employee_id)
    if not data.get("employee"):
        return Response("Karyawan tidak ditemukan.", status=404)
    latest = data["submissions"][0] if data["submissions"] else None
    insights = build_insights(latest, current_year=datetime.now().year) if latest else None
    pdf_bytes = pdf_export.employee_pdf(data["employee"], data["submissions"], latest, insights)
    nama = data["employee"].get("full_name") or f"id{employee_id}"
    fname = f"karyawan-{_ascii_filename(nama, 'karyawan')}.pdf"
    return Response(pdf_bytes, mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ---------------------------------------------------------------------------
# Export XLSX (data terbaru per laptop) — anti formula injection.
# ---------------------------------------------------------------------------
_EXPORT_COLUMNS = [
    # — Identitas pengisian & pemegang —
    ("submitted_at",        "Waktu Pengisian"),
    ("source",              "Sumber Data"),
    ("holder_name",         "Nama Karyawan"),
    ("holder_position",     "Jabatan"),
    ("holder_company",      "Perusahaan"),
    ("holder_location",     "Lokasi Penempatan"),
    ("work_group",          "Kelompok Kerja"),
    ("laptop_status",       "Status Kepemilikan"),
    # — Identitas aset —
    ("serial_number",       "Nomor Seri"),
    ("asset_no",            "No. Aset"),
    ("hostname",            "Hostname"),
    ("mac_address",         "MAC Address"),
    ("device_brand",        "Merek"),
    ("device_model",        "Model Laptop"),
    # — CPU —
    ("cpu_model",           "CPU / Prosesor"),
    ("cpu_passmark",        "Skor CPU (PassMark)"),
    ("cpu_cores",           "Core CPU"),
    ("cpu_threads",         "Thread CPU"),
    ("cpu_arch",            "Arsitektur CPU"),
    ("cpu_speed_mhz",       "Kecepatan CPU (MHz)"),
    ("cpu_usage_pct",       "Pemakaian CPU (%)"),
    # — GPU —
    ("gpu",                 "GPU / Kartu Grafis"),
    # — RAM —
    ("ram_gb",              "RAM (GB)"),
    ("ram_type",            "Tipe RAM"),
    ("ram_speed_mhz",       "Kecepatan RAM (MHz)"),
    ("ram_usage_pct",       "Pemakaian RAM (%)"),
    ("ram_usage_gb",        "RAM Terpakai (GB)"),
    # — Motherboard & RAM (board) —
    ("motherboard",         "Motherboard"),
    ("ram_slots_used",      "Slot RAM Terisi"),
    ("ram_slots_total",     "Slot RAM Total"),
    ("ram_max_gb",          "RAM Maksimum (GB)"),
    # — Penyimpanan & OS —
    ("ssd_gb",              "SSD (GB)"),
    ("ssd_type",            "Tipe SSD"),
    ("hdd_gb",              "HDD (GB)"),
    ("disk_health_pct",     "Kesehatan Disk (%)"),
    ("disk_health_raw",     "Status Disk"),
    ("os_total_gb",         "Kapasitas Disk OS (GB)"),
    ("os_free_gb",          "Sisa Disk OS (GB)"),
    ("os_name",             "Sistem Operasi"),
    # — Keamanan & kesiapan Windows 11 —
    ("tpm_version",         "Versi TPM"),
    ("secure_boot",         "Secure Boot"),
    ("win11_ready",         "Siap Windows 11"),
    ("win11_blockers",      "Kendala Windows 11"),
    # — Baterai —
    ("battery_pct",         "Daya Baterai saat Pengecekan (%)"),
    ("battery_health_pct",  "Kesehatan Baterai (%)"),
    ("battery_wh_full",     "Kapasitas Penuh Baterai (Wh)"),
    ("battery_wh_design",   "Kapasitas Desain Baterai (Wh)"),
    # — Kondisi & kelengkapan —
    ("physical_condition",  "Kondisi Fisik"),
    ("accessories",         "Kelengkapan"),
    ("purchase_year",       "Tahun Pembelian"),
    ("issues",              "Kerusakan / Keluhan"),
    # — Hasil penilaian —
    ("score_spec",          "Skor Spek"),
    ("score_load",          "Skor Beban"),
    ("score_total",         "Skor Total"),
    ("status",              "Status Kelayakan"),
    ("eol_year",            "Estimasi Pensiun"),
    ("status_reasons",      "Alasan Status"),
    # — Insight (turunan) —
    ("_verdict",            "Kesimpulan"),
    ("_two_axis",           "Spek vs Beban Nyata"),
    ("_recommendations",    "Rekomendasi"),
    ("_placement",          "Saran Penempatan"),
]


def _battery_health_value(r):
    """Kesehatan baterai (%): pakai kolom tersimpan, else hitung dari Wh."""
    h = r.get("battery_health_pct")
    if h is not None:
        return round(h)
    full, design = r.get("battery_wh_full"), r.get("battery_wh_design")
    if full and design and design > 0:
        return round(full / design * 100)
    return None


def _sanitize(val):
    """Cegah CSV/Excel formula injection: prefix ' bila diawali = + - @."""
    if val is None:
        return ""
    s = str(val)
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _yes_no_dash(v):
    """Map flag 1/0/None -> 'Ya'/'Tidak'/'-' untuk ekspor."""
    if v in (1, "1"):
        return "Ya"
    if v in (0, "0"):
        return "Tidak"
    return "-"


# Kolom sheet "Per Karyawan" (key pada baris latest_per_employee()).
_EMPLOYEE_COLUMNS = [
    ("emp_full_name",          "Nama"),
    ("emp_company",            "Perusahaan"),
    ("emp_current_position",   "Jabatan"),
    ("_group",                 "Kelompok"),
    ("_laptop",                "Laptop"),
    ("device_serial",          "Serial"),
    ("score_total",            "Skor"),
    ("_status",                "Status"),
    ("_employee_status",       "Status Karyawan"),
    ("submitted_at",           "Terakhir"),
]


def _write_employee_sheet(wb, header_font, header_fill):
    """Tambah sheet 'Per Karyawan' dari db.latest_per_employee()."""
    try:
        emps = db.latest_per_employee()
    except AttributeError:
        emps = []

    ws = wb.create_sheet("Per Karyawan")
    for col, (_key, label) in enumerate(_EMPLOYEE_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill

    for ri, r in enumerate(emps, 2):
        for ci, (key, _label) in enumerate(_EMPLOYEE_COLUMNS, 1):
            if key == "_group":
                # Pakai kelompok karyawan terkini; fallback ke work_group submission.
                wg = r.get("emp_current_work_group") or r.get("work_group")
                value = wg_labels().get(wg, group_label(r))
            elif key == "_laptop":
                value = ((r.get("device_brand") or "") + " "
                         + (r.get("device_model") or "")).strip() or "-"
            elif key == "_status":
                value = STATUS_LABEL.get(r.get("status"), r.get("status"))
            elif key == "_employee_status":
                value = "Aktif" if (r.get("emp_status") or "active") == "active" else "Resign"
            else:
                value = r.get(key)
            ws.cell(row=ri, column=ci, value=_sanitize(value))

    for col, (_k, label) in enumerate(_EMPLOYEE_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(12, len(label) + 2)
    ws.freeze_panes = "A2"


@admin_bp.route("/export.xlsx")
@login_required
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = db.latest_per_device()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaris Laptop"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col, (_key, label) in enumerate(_EXPORT_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill

    for ri, r in enumerate(rows, 2):
        insight = build_insights(r)
        placement = suggest_placement(r)
        for ci, (key, _label) in enumerate(_EXPORT_COLUMNS, 1):
            if key == "status_reasons":
                value = "; ".join(_parse_reasons(r.get("status_reasons")))
            elif key == "work_group":
                value = group_label(r)
            elif key == "status":
                value = STATUS_LABEL.get(r.get("status"), r.get("status"))
            elif key == "laptop_status":
                value = LAPTOP_STATUS_LABEL.get(r.get("laptop_status"), r.get("laptop_status"))
            elif key == "physical_condition":
                value = PHYSICAL_CONDITION_LABEL.get(r.get("physical_condition"), r.get("physical_condition"))
            elif key == "source":
                value = SOURCE_LABEL.get(r.get("source"), r.get("source"))
            elif key == "battery_health_pct":
                value = _battery_health_value(r)
            elif key == "secure_boot":
                value = _yes_no_dash(r.get("secure_boot"))
            elif key == "win11_ready":
                value = _yes_no_dash(r.get("win11_ready"))
            elif key == "_verdict":
                value = insight["verdict"]
            elif key == "_two_axis":
                ta = insight.get("two_axis")
                value = ta["headline"] if ta else None
            elif key == "_recommendations":
                value = " | ".join(insight["recommendations"])
            elif key == "_placement":
                value = (placement.get("text") if placement else None) or "Sesuai perannya"
            else:
                value = r.get(key)
            ws.cell(row=ri, column=ci, value=_sanitize(value))

    # Lebar kolom sederhana.
    for col, (_k, label) in enumerate(_EXPORT_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(12, len(label) + 2)
    ws.freeze_panes = "A2"

    # — Sheet kedua: Per Karyawan —
    _write_employee_sheet(wb, header_font, header_fill)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"inventaris-laptop-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
