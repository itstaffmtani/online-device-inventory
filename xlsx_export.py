# xlsx_export.py — Export XLSX 3+1 sheet, skoring TRANSPARAN lewat rumus Excel.
#
# Latar: ekspor lama = 1 sheet "bahasa manusia" (label jadi, tak bisa ditelusuri).
# Permintaan: pisahkan PARAMETER (master) dari DATA, dan biar Skor dihitung oleh
# RUMUS di dalam Excel (referensi ke master) sehingga bisa diaudit & dikalibrasi
# langsung di spreadsheet. Yang TIDAK mungkin dirumuskan di Excel = PassMark CPU
# (butuh tabel benchmark offline) -> tetap nilai masukan apa adanya.
#
# Struktur:
#   Sheet 1 "Master"      — parameter per kelompok (cpu/ram floor·ideal, bobot) +
#                           pengaturan global (ambang, blend, masa pakai) + kamus
#                           kode->label. SATU sumber angka; rumus menunjuk ke sini.
#   Sheet 2 "Perhitungan" — 1 baris/laptop: masukan mentah + kolom turunan & SKOR
#                           via rumus (VLOOKUP ke Master). Inilah mesin yang bisa
#                           ditelusuri sel demi sel. Mengikuti scoring.py persis.
#   Sheet 3 "Ringkasan"   — bahasa manusia: lookup dari Master + Perhitungan
#                           (label kelompok/status, skor). PassMark dikutip langsung.
#   Sheet 4 "Per Karyawan"— satu baris per karyawan (laptop terbarunya).
#
# Catatan paritas dgn scoring.py (sumber kebenaran):
#   spec  = rata2 berbobot {cpu,ram,storage(+disk health),battery}; baterai keluar
#           dari bobot bila datanya kosong.
#   load  = 0.5*tekanan(ramp linier 60->100%) + 0.5*kecukupan(RAM vs min); -10 bila
#           sisa disk OS < 15%.
#   total = blend_spec*spec + blend_load*load.
#   status= ambang(total) lalu DITURUNKAN ke Upgrade bila RAM<min / HDD-tanpa-SSD /
#           CPU<floor (tak pernah menaikkan; "Ganti" hanya dari skor rendah).

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_SUB_FILL = PatternFill("solid", fgColor="EEF2FF")
_SCORE_FILL = PatternFill("solid", fgColor="FEF3C7")
_TITLE_FONT = Font(bold=True, size=12, color="3730A3")
_LABEL_FONT = Font(bold=True, color="334155")


def _sanitize(val):
    """Cegah CSV/Excel formula injection untuk teks: prefiks ' bila diawali = + - @."""
    if val is None:
        return None
    s = str(val)
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _num(v):
    """Angka -> float/int; None bila kosong/tak terbaca (sel dibiarkan KOSONG)."""
    if v is None or v == "":
        return None
    try:
        f = float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None
    return int(f) if f == int(f) else f


# ---------------------------------------------------------------------------
# Sheet 1 — Master (parameter + pengaturan + kamus)
# ---------------------------------------------------------------------------
# Kolom tabel kelompok (A..K). Indeks VLOOKUP = posisi di sini (1-based).
_GROUP_COLS = [
    ("key", "Kode"), ("label", "Kelompok"), ("profile_label", "Profil"),
    ("cpu_floor", "CPU min"), ("cpu_ideal", "CPU ideal"),
    ("ram_min", "RAM min"), ("ram_ideal", "RAM ideal"),
    ("w_cpu", "Bobot CPU"), ("w_ram", "Bobot RAM"),
    ("w_storage", "Bobot Storage"), ("w_battery", "Bobot Baterai"),
]
_GCOL = {name: i + 1 for i, (name, _l) in enumerate(_GROUP_COLS)}  # 1-based
_GROUP_HEADER_ROW = 3
_GROUP_FIRST_ROW = 4

# Pengaturan global di kolom M(13)/N(14).
_SET_COL_LABEL = 13   # M
_SET_COL_VALUE = 14   # N
# Kamus di kolom P(16)/Q(17).
_KAM_COL_CODE = 16    # P
_KAM_COL_LABEL = 17   # Q


def _build_master(wb, groups, settings):
    """Sheet Master. Kembalikan dict referensi sel absolut untuk dipakai rumus."""
    ws = wb.active
    ws.title = "Master"

    ws.cell(row=1, column=1, value="PARAMETER SKORING — sumber angka (edit di sini, skor ikut berubah)").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "Catatan: skor di sheet 'Perhitungan' dihitung ulang oleh rumus Excel "
        "mengikuti aturan aplikasi. Bisa berbeda ≤1 poin dari aplikasi hanya pada "
        "pembulatan tepat .5 (Excel membulatkan ke atas).")).font = Font(italic=True, color="64748B")

    # — Tabel kelompok —
    for ci, (_name, label) in enumerate(_GROUP_COLS, 1):
        c = ws.cell(row=_GROUP_HEADER_ROW, column=ci, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
    r = _GROUP_FIRST_ROW
    for g in groups:
        for ci, (name, _l) in enumerate(_GROUP_COLS, 1):
            ws.cell(row=r, column=ci, value=_sanitize(g.get(name)) if name in
                    ("key", "label", "profile_label") else _num(g.get(name)))
        r += 1
    last_group_row = r - 1
    groups_range = f"Master!$A${_GROUP_FIRST_ROW}:${get_column_letter(len(_GROUP_COLS))}${last_group_row}"

    L, V = get_column_letter(_SET_COL_LABEL), get_column_letter(_SET_COL_VALUE)

    def _set(row, label, value):
        ws.cell(row=row, column=_SET_COL_LABEL, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=_SET_COL_VALUE, value=value)
        return f"Master!${V}${row}"

    ws.cell(row=_GROUP_HEADER_ROW, column=_SET_COL_LABEL,
            value="PENGATURAN GLOBAL").font = _HEADER_FONT
    refs = {}
    refs["eligible"] = _set(4, "Ambang Layak (≥)", _num(settings.get("status_eligible_min")))
    refs["upgrade"] = _set(5, "Ambang Upgrade (≥)", _num(settings.get("status_upgrade_min")))
    refs["blend_spec"] = _set(6, "Bobot Spek (blend)", _num(settings.get("blend_spec")))
    refs["blend_load"] = _set(7, "Bobot Beban (blend)", _num(settings.get("blend_load")))
    refs["lifespan"] = _set(8, "Masa Pakai (tahun)", _num(settings.get("base_lifespan_years")))

    ws.cell(row=10, column=_SET_COL_LABEL, value="STORAGE (poin dasar)").font = _HEADER_FONT
    refs["sto_nvme"] = _set(11, "SSD NVMe", 100)
    refs["sto_ssd"] = _set(12, "SSD SATA", 85)
    refs["sto_hdd"] = _set(13, "HDD saja", 25)
    refs["sto_unk"] = _set(14, "Tak terdeteksi", 50)

    # — Kamus status (kode -> label) —
    kc, kl = get_column_letter(_KAM_COL_CODE), get_column_letter(_KAM_COL_LABEL)
    ws.cell(row=_GROUP_HEADER_ROW, column=_KAM_COL_CODE, value="KAMUS STATUS").font = _HEADER_FONT
    status_map = [("eligible", "Layak"), ("upgrade", "Upgrade"), ("replace", "Ganti")]
    for i, (code, label) in enumerate(status_map):
        ws.cell(row=4 + i, column=_KAM_COL_CODE, value=code)
        ws.cell(row=4 + i, column=_KAM_COL_LABEL, value=label)
    refs["status_kamus"] = f"Master!${kc}$4:${kl}${4 + len(status_map) - 1}"

    ws.cell(row=8, column=_KAM_COL_CODE, value="KAMUS KEPEMILIKAN").font = _HEADER_FONT
    own_map = [("office_inventory", "Inventaris Kantor"), ("personal", "Milik Pribadi")]
    for i, (code, label) in enumerate(own_map):
        ws.cell(row=9 + i, column=_KAM_COL_CODE, value=code)
        ws.cell(row=9 + i, column=_KAM_COL_LABEL, value=label)
    refs["own_kamus"] = f"Master!${kc}$9:${kl}${9 + len(own_map) - 1}"

    refs["groups_range"] = groups_range
    # Lebar kolom.
    for ci in range(1, len(_GROUP_COLS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.column_dimensions[L].width = 20
    ws.column_dimensions[V].width = 10
    ws.column_dimensions[kc].width = 16
    ws.column_dimensions[kl].width = 18
    ws.freeze_panes = "A4"
    return refs


# ---------------------------------------------------------------------------
# Sheet 2 — Perhitungan (rumus, paritas scoring.py)
# ---------------------------------------------------------------------------
# (field, header, kind). kind: text|num|formula.
_CALC_COLS = [
    ("submitted_at", "Waktu", "text"),
    ("holder_name", "Nama", "text"),
    ("laptop", "Laptop", "text"),
    ("serial", "Serial", "text"),
    ("work_group", "Kode Kelompok", "text"),
    ("grp_eff", "Kelompok Efektif", "formula"),
    ("cpu_passmark", "PassMark", "num"),
    ("ram_gb", "RAM (GB)", "num"),
    ("ssd_gb", "SSD (GB)", "num"),
    ("ssd_type", "Tipe SSD", "text"),
    ("hdd_gb", "HDD (GB)", "num"),
    ("disk_health_pct", "Sehat Disk (%)", "num"),
    ("battery_health_pct", "Sehat Baterai (%)", "num"),
    ("battery_wh_full", "Baterai Penuh (Wh)", "num"),
    ("battery_wh_design", "Baterai Desain (Wh)", "num"),
    ("ram_usage_pct", "Pakai RAM (%)", "num"),
    ("cpu_usage_pct", "Pakai CPU (%)", "num"),
    ("os_total_gb", "Disk OS Total (GB)", "num"),
    ("os_free_gb", "Disk OS Sisa (GB)", "num"),
    ("purchase_year", "Tahun Beli", "num"),
    ("laptop_status", "Kode Kepemilikan", "text"),
    ("cpu_floor", "CPU min", "formula"),
    ("cpu_ideal", "CPU ideal", "formula"),
    ("ram_min", "RAM min", "formula"),
    ("ram_ideal", "RAM ideal", "formula"),
    ("w_cpu", "Bobot CPU", "formula"),
    ("w_ram", "Bobot RAM", "formula"),
    ("w_storage", "Bobot Storage", "formula"),
    ("w_battery", "Bobot Baterai", "formula"),
    ("cpu_pts", "Poin CPU", "formula"),
    ("ram_pts", "Poin RAM", "formula"),
    ("sto_base", "Poin Storage dasar", "formula"),
    ("sto_pts", "Poin Storage", "formula"),
    ("bat_health", "Kesehatan Baterai (%)", "formula"),
    ("bat_pts", "Poin Baterai", "formula"),
    ("spec", "SKOR SPEK", "formula"),
    ("avg_pct", "Rata Beban (%)", "formula"),
    ("pressure", "Tekanan", "formula"),
    ("adequacy", "Kecukupan", "formula"),
    ("load_base", "Beban dasar", "formula"),
    ("os_low", "Disk OS sesak?", "formula"),
    ("load", "SKOR BEBAN", "formula"),
    ("total", "SKOR TOTAL", "formula"),
    ("status_base", "Status (ambang)", "formula"),
    ("force", "Paksa upgrade?", "formula"),
    ("status", "STATUS", "formula"),
    ("eol", "Est. Pensiun", "formula"),
]
_CCOL = {f: i + 1 for i, (f, _h, _k) in enumerate(_CALC_COLS)}


def _cl(field):
    """Huruf kolom untuk field di sheet Perhitungan."""
    return get_column_letter(_CCOL[field])


def _calc_formula(field, r, refs):
    """Rumus Excel untuk `field` di baris r. Menunjuk ke Master via `refs`."""
    c = lambda f: f"{_cl(f)}{r}"  # alamat sel field lain di baris yang sama
    g = refs["groups_range"]
    key = c("work_group")
    grp = c("grp_eff")

    def vlk(idx):  # VLOOKUP parameter kelompok efektif -> kolom idx tabel groups
        return f"VLOOKUP({grp},{g},{idx},FALSE)"

    has_ssd = f'OR({c("ssd_gb")}>0,TRIM({c("ssd_type")})<>"")'

    if field == "grp_eff":
        return f'=IFERROR(VLOOKUP({key},{g},1,FALSE),"admin")'
    if field == "cpu_floor":
        return f"={vlk(_GCOL['cpu_floor'])}"
    if field == "cpu_ideal":
        return f"={vlk(_GCOL['cpu_ideal'])}"
    if field == "ram_min":
        return f"={vlk(_GCOL['ram_min'])}"
    if field == "ram_ideal":
        return f"={vlk(_GCOL['ram_ideal'])}"
    if field == "w_cpu":
        return f"={vlk(_GCOL['w_cpu'])}"
    if field == "w_ram":
        return f"={vlk(_GCOL['w_ram'])}"
    if field == "w_storage":
        return f"={vlk(_GCOL['w_storage'])}"
    if field == "w_battery":
        return f"={vlk(_GCOL['w_battery'])}"
    if field == "cpu_pts":
        return f'=MIN(100,MAX(0,ROUND(100*{c("cpu_passmark")}/{c("cpu_ideal")},0)))'
    if field == "ram_pts":
        return f'=MIN(100,MAX(0,ROUND(100*{c("ram_gb")}/{c("ram_ideal")},0)))'
    if field == "sto_base":
        return (f'=IF({has_ssd},IF(ISNUMBER(SEARCH("nvme",{c("ssd_type")})),'
                f'{refs["sto_nvme"]},{refs["sto_ssd"]}),'
                f'IF({c("hdd_gb")}>0,{refs["sto_hdd"]},{refs["sto_unk"]}))')
    if field == "sto_pts":
        dh = c("disk_health_pct")
        return (f'=IF({dh}="",{c("sto_base")},'
                f'ROUND({c("sto_base")}*MIN(1,MAX(0.5,{dh}/100)),0))')
    if field == "bat_health":
        bh, full, des = c("battery_health_pct"), c("battery_wh_full"), c("battery_wh_design")
        return (f'=IF({bh}<>"",{bh},'
                f'IF(AND({full}<>"",{des}<>"",{des}>0),{full}/{des}*100,""))')
    if field == "bat_pts":
        bh = c("bat_health")
        return f'=IF({bh}="","",MIN(100,MAX(0,ROUND({bh},0))))'
    if field == "spec":
        bh = c("bat_health")
        num = (f'{c("cpu_pts")}*{c("w_cpu")}+{c("ram_pts")}*{c("w_ram")}'
               f'+{c("sto_pts")}*{c("w_storage")}'
               f'+IF({bh}<>"",{c("bat_pts")}*{c("w_battery")},0)')
        den = f'{c("w_cpu")}+{c("w_ram")}+{c("w_storage")}+IF({bh}<>"",{c("w_battery")},0)'
        return f"=ROUND(({num})/({den}),0)"
    if field == "avg_pct":
        rp, cp = c("ram_usage_pct"), c("cpu_usage_pct")
        return f'=IF(COUNT({rp},{cp})=0,"",AVERAGE({rp},{cp}))'
    if field == "pressure":
        a = c("avg_pct")
        return (f'=IF({a}="",100,IF({a}<=60,100,IF({a}>=100,0,'
                f'100*(100-{a})/40)))')
    if field == "adequacy":
        return f'=MIN(100,MAX(0,ROUND(100*{c("ram_gb")}/{c("ram_min")},0)))'
    if field == "load_base":
        return f'=ROUND(0.5*{c("pressure")}+0.5*{c("adequacy")},0)'
    if field == "os_low":
        ot, of = c("os_total_gb"), c("os_free_gb")
        return f'=IF(AND({ot}<>"",{ot}>0,{of}<>""),{of}/{ot}<0.15,FALSE)'
    if field == "load":
        return f'=IF({c("os_low")},MAX(0,{c("load_base")}-10),{c("load_base")})'
    if field == "total":
        return f'=ROUND({refs["blend_spec"]}*{c("spec")}+{refs["blend_load"]}*{c("load")},0)'
    if field == "status_base":
        t = c("total")
        return (f'=IF({t}>={refs["eligible"]},"eligible",'
                f'IF({t}>={refs["upgrade"]},"upgrade","replace"))')
    if field == "force":
        pm, floor = c("cpu_passmark"), c("cpu_floor")
        ram, rmin = c("ram_gb"), c("ram_min")
        hdd = c("hdd_gb")
        return (f'=OR(AND({ram}>0,{ram}<{rmin}),'
                f'AND(NOT({has_ssd}),{hdd}>0),'
                f'AND({pm}>0,{pm}<{floor}))')
    if field == "status":
        return (f'=IF(AND({c("force")},{c("status_base")}="eligible"),'
                f'"upgrade",{c("status_base")})')
    if field == "eol":
        py, ls = c("purchase_year"), c("laptop_status")
        return f'=IF(OR(LOWER({ls})="personal",{py}=""),"",{py}+{refs["lifespan"]})'
    return ""


def _calc_value(field, row):
    """Nilai masukan mentah (non-formula) untuk `field`."""
    if field == "laptop":
        return _sanitize(((row.get("device_brand") or "") + " "
                          + (row.get("device_model") or "")).strip() or None)
    if field == "serial":
        return _sanitize(row.get("serial_number") or row.get("device_serial"))
    if field == "submitted_at":
        return _sanitize(row.get("submitted_at"))
    kind = next(k for f, _h, k in _CALC_COLS if f == field)
    if kind == "num":
        return _num(row.get(field))
    return _sanitize(row.get(field))


def _build_calc(wb, rows, refs):
    """Sheet Perhitungan. Kembalikan jumlah baris data."""
    ws = wb.create_sheet("Perhitungan")
    for ci, (_f, header, _k) in enumerate(_CALC_COLS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for ri, row in enumerate(rows, 2):
        for ci, (field, _h, kind) in enumerate(_CALC_COLS, 1):
            cell = ws.cell(row=ri, column=ci)
            if kind == "formula":
                cell.value = _calc_formula(field, ri, refs)
                if field in ("spec", "load", "total"):
                    cell.fill = _SCORE_FILL
                    cell.font = Font(bold=True)
            else:
                cell.value = _calc_value(field, row)

    for ci in range(1, len(_CALC_COLS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "F2"  # kunci kolom identitas + judul
    return len(rows)


# ---------------------------------------------------------------------------
# Sheet 3 — Ringkasan (bahasa manusia; lookup Master + Perhitungan)
# ---------------------------------------------------------------------------
_SUMMARY_COLS = [
    ("Nama", "holder_name"), ("Laptop", "laptop"), ("Serial", "serial"),
    ("Kelompok", "group_label"), ("CPU", "cpu_model"), ("PassMark", "cpu_passmark"),
    ("RAM (GB)", "ram_gb"), ("Penyimpanan", "storage_txt"),
    ("Skor Spek", "spec"), ("Skor Beban", "load"), ("Skor Total", "total"),
    ("Status", "status_label"), ("Est. Pensiun", "eol"),
]


def _build_summary(wb, rows, refs, n_calc):
    """Sheet Ringkasan: tiap sel skor/label MENUNJUK ke Perhitungan/Master (lookup),
    kecuali PassMark & CPU yang dikutip langsung (tak bisa dirumus)."""
    ws = wb.create_sheet("Ringkasan")
    for ci, (header, _f) in enumerate(_SUMMARY_COLS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL

    for i, row in enumerate(rows):
        cr = i + 2              # baris di sheet Perhitungan
        sr = i + 2              # baris di sheet Ringkasan
        for ci, (_header, field) in enumerate(_SUMMARY_COLS, 1):
            cell = ws.cell(row=sr, column=ci)
            if field == "laptop":
                cell.value = _sanitize(((row.get("device_brand") or "") + " "
                                        + (row.get("device_model") or "")).strip() or None)
            elif field == "serial":
                cell.value = _sanitize(row.get("serial_number") or row.get("device_serial"))
            elif field == "cpu_model":
                cell.value = _sanitize(row.get("cpu_model"))
            elif field == "cpu_passmark":
                cell.value = _num(row.get("cpu_passmark"))
            elif field == "ram_gb":
                cell.value = _num(row.get("ram_gb"))
            elif field == "storage_txt":
                cell.value = _sanitize(_storage_text(row))
            elif field == "holder_name":
                cell.value = _sanitize(row.get("holder_name"))
            elif field == "group_label":
                # lookup kode kelompok -> label di tabel Master.
                cell.value = (f"=IFERROR(VLOOKUP(Perhitungan!{_cl('work_group')}{cr},"
                              f"{refs['groups_range']},{_GCOL['label']},FALSE),"
                              f"Perhitungan!{_cl('work_group')}{cr})")
            elif field == "spec":
                cell.value = f"=Perhitungan!{_cl('spec')}{cr}"
            elif field == "load":
                cell.value = f"=Perhitungan!{_cl('load')}{cr}"
            elif field == "total":
                cell.value = f"=Perhitungan!{_cl('total')}{cr}"
            elif field == "status_label":
                cell.value = (f"=IFERROR(VLOOKUP(Perhitungan!{_cl('status')}{cr},"
                              f"{refs['status_kamus']},2,FALSE),"
                              f"Perhitungan!{_cl('status')}{cr})")
            elif field == "eol":
                cell.value = f"=Perhitungan!{_cl('eol')}{cr}"

    widths = [22, 26, 16, 16, 26, 10, 9, 22, 10, 10, 10, 12, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def _storage_text(row):
    ssd = _num(row.get("ssd_gb"))
    hdd = _num(row.get("hdd_gb"))
    typ = (row.get("ssd_type") or "").strip()
    parts = []
    if ssd:
        parts.append(f"{ssd} GB SSD" + (f" ({typ})" if typ else ""))
    if hdd:
        parts.append(f"{hdd} GB HDD")
    return " + ".join(parts) or "—"


# ---------------------------------------------------------------------------
# Sheet 4 — Per Karyawan
# ---------------------------------------------------------------------------
_EMPLOYEE_COLS = [
    ("Nama", "emp_full_name"), ("Perusahaan", "emp_company"),
    ("Jabatan", "emp_current_position"), ("Kelompok", "_group"),
    ("Laptop", "_laptop"), ("Serial", "device_serial"),
    ("Skor", "score_total"), ("Status", "_status"),
    ("Status Karyawan", "_emp_status"), ("Terakhir", "submitted_at"),
]


def _build_employees(wb, employees, wg_labels, status_label):
    ws = wb.create_sheet("Per Karyawan")
    for ci, (header, _f) in enumerate(_EMPLOYEE_COLS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
    for ri, r in enumerate(employees, 2):
        for ci, (_header, field) in enumerate(_EMPLOYEE_COLS, 1):
            if field == "_group":
                wg = r.get("emp_current_work_group") or r.get("work_group")
                value = wg_labels.get(wg, wg or "-")
            elif field == "_laptop":
                value = ((r.get("device_brand") or "") + " "
                         + (r.get("device_model") or "")).strip() or "-"
            elif field == "_status":
                value = status_label.get(r.get("status"), r.get("status"))
            elif field == "_emp_status":
                value = "Aktif" if (r.get("emp_status") or "active") == "active" else "Resign"
            elif field == "score_total":
                value = _num(r.get("score_total"))
            else:
                value = r.get(field)
            ws.cell(row=ri, column=ci,
                    value=value if isinstance(value, (int, float)) else _sanitize(value))
    for ci, (header, _f) in enumerate(_EMPLOYEE_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(header) + 2)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# API utama
# ---------------------------------------------------------------------------
def build_workbook(rows, groups, settings, employees, wg_labels, status_label):
    """Rakit Workbook 4 sheet (Master, Perhitungan, Ringkasan, Per Karyawan)."""
    wb = Workbook()
    refs = _build_master(wb, groups, settings)
    n = _build_calc(wb, rows, refs)
    _build_summary(wb, rows, refs, n)
    _build_employees(wb, employees, wg_labels, status_label)
    return wb
